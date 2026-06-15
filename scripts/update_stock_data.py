from __future__ import annotations

import argparse
import json
import logging
import math
import shutil
import sys
import urllib.parse
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
import socket
from zoneinfo import ZoneInfo

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import yaml
except ModuleNotFoundError:
    yaml = None

try:
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry
except ModuleNotFoundError:
    HTTPAdapter = None
    Retry = None


ROOT = Path(__file__).resolve().parents[1]


@dataclass
class Holding:
    row_number: int
    tracking: str
    market: str
    stock_id: str
    stock_name: str
    category: str
    shares: float | None
    average_cost: float | None
    buy_date: Any
    purpose: str
    rule: str
    note: str


def load_config() -> dict[str, Any]:
    config_path = ROOT / "config" / "config.yaml"
    if yaml is None:
        return {
            "timezone": "Asia/Taipei",
            "run_time": "16:00",
            "input_excel": "input/watchlist.xlsx",
            "manual_sheet": "01_手動填寫持股",
            "output_sheet": "04_自動更新結果",
            "source_sheet": "09_資料來源",
            "reports_folder": "reports",
            "logs_folder": "logs",
            "raw_data_folder": "data_raw",
            "clean_data_folder": "data_clean",
            "history_calendar_days": 420,
            "fundamental_calendar_days": 900,
            "chip_calendar_days": 45,
            "news_calendar_days": 14,
            "request_timeout_seconds": 20,
            "api": {
                "twse_stock_day_all": "https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL",
                "finmind_data": "https://api.finmindtrade.com/api/v4/data",
            },
            "manual_columns": {
                "tracking": "是否追蹤(Y/N)",
                "market": "市場(TWSE/TPEx/US)",
                "stock_id": "股票代號",
                "stock_name": "股票名稱",
                "category": "投資分類",
                "shares": "持有股數",
                "average_cost": "成本均價",
                "buy_date": "買進日期",
                "purpose": "投資目的",
                "rule": "停損/減碼規則",
                "note": "備註",
            },
            "stage2_sheets": {
                "monthly_revenue": "11_月營收追蹤",
                "institutional": "12_法人籌碼追蹤",
                "financial": "13_財報估值追蹤",
                "news": "14_新聞風險追蹤",
                "score": "15_第二階段綜合評分",
            },
            "public_dashboard": {
                "publish_position_amounts": False,
                "publish_shares": False,
                "publish_average_cost": False,
                "publish_raw_source_paths": False,
            },
        }
    with config_path.open("r", encoding="utf-8") as fh:
        return yaml.safe_load(fh)


def load_public_view_settings() -> dict[str, Any]:
    path = ROOT / "config" / "public_view_settings.json"
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return {
        "viewMode": "family",
        "showPnlRate": True,
        "showCostBasis": False,
        "showShares": False,
        "showMarketValue": False,
        "showUnrealizedPnlAmount": False,
        "showTotalAssets": False,
    }


def ensure_dirs(config: dict[str, Any]) -> None:
    for key in ["reports_folder", "logs_folder", "raw_data_folder", "clean_data_folder"]:
        (ROOT / config[key]).mkdir(parents=True, exist_ok=True)


def setup_logger(log_path: Path) -> logging.Logger:
    logger = logging.getLogger("stock_tracker")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setFormatter(fmt)
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(fmt)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    return logger


def norm_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        if isinstance(value, str):
            value = value.replace(",", "").strip()
        return float(value)
    except (TypeError, ValueError):
        return None


def read_holdings(input_excel: Path, manual_sheet: str, column_names: dict[str, str]) -> list[Holding]:
    wb = load_workbook(input_excel, data_only=False)
    if manual_sheet not in wb.sheetnames:
        raise ValueError(f"找不到工作表：{manual_sheet}")
    ws = wb[manual_sheet]
    headers = {norm_text(ws.cell(2, col).value): col for col in range(1, ws.max_column + 1)}

    def col(logical_name: str) -> int:
        wanted = column_names[logical_name]
        if wanted not in headers:
            raise ValueError(f"手動表缺少欄位：{wanted}")
        return headers[wanted]

    holdings: list[Holding] = []
    for row in range(3, ws.max_row + 1):
        stock_id = norm_text(ws.cell(row, col("stock_id")).value)
        if not stock_id:
            continue
        holdings.append(
            Holding(
                row_number=row,
                tracking=norm_text(ws.cell(row, col("tracking")).value).upper(),
                market=norm_text(ws.cell(row, col("market")).value).upper(),
                stock_id=stock_id.zfill(4) if stock_id.isdigit() else stock_id,
                stock_name=norm_text(ws.cell(row, col("stock_name")).value),
                category=norm_text(ws.cell(row, col("category")).value),
                shares=to_float(ws.cell(row, col("shares")).value),
                average_cost=to_float(ws.cell(row, col("average_cost")).value),
                buy_date=ws.cell(row, col("buy_date")).value,
                purpose=norm_text(ws.cell(row, col("purpose")).value),
                rule=norm_text(ws.cell(row, col("rule")).value),
                note=norm_text(ws.cell(row, col("note")).value),
            )
        )
    return holdings


def request_json(url: str, params: dict[str, Any] | None, timeout: int, logger: logging.Logger) -> Any:
    logger.info("API request url=%s params=%s", url, params or {})
    session = requests.Session()
    if HTTPAdapter and Retry:
        retry = Retry(
            total=3,
            connect=3,
            read=3,
            status=3,
            backoff_factor=1.5,
            status_forcelist=(429, 500, 502, 503, 504),
            allowed_methods=frozenset(["GET"]),
            raise_on_status=False,
        )
        adapter = HTTPAdapter(max_retries=retry)
        session.mount("http://", adapter)
        session.mount("https://", adapter)
    headers = {
        "User-Agent": "stock-tracker-v1 (+https://github.com/leon5281ss/taylorstock-dashboard)",
        "Accept": "application/json,text/plain,*/*",
    }
    try:
        resp = session.get(url, params=params, headers=headers, timeout=timeout)
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.Timeout as exc:
        raise RuntimeError(f"HTTPS timeout: {exc}") from exc
    except requests.exceptions.ConnectionError as exc:
        raise RuntimeError(f"HTTPS connection failed: {exc}") from exc
    except requests.exceptions.HTTPError as exc:
        code = exc.response.status_code if exc.response is not None else "unknown"
        raise RuntimeError(f"HTTP {code}: {exc}") from exc
    except requests.exceptions.RequestException as exc:
        raise RuntimeError(f"Request failed: {exc}") from exc
    finally:
        session.close()


def classify_api_error(exc: Exception) -> str:
    message = str(exc)
    if isinstance(exc, socket.gaierror) or "Name or service not known" in message or "getaddrinfo failed" in message:
        return "DNS 失敗"
    if "HTTPS timeout" in message or "timed out" in message:
        return "HTTPS 連線失敗"
    if "403" in message or "401" in message or "Unauthorized" in message or "Forbidden" in message:
        return "API 權限不足"
    if "429" in message or "Too Many Requests" in message:
        return "API 限流"
    if "not found" in message.lower() or "無效資料" in message:
        return "資料格式改變"
    if "does not exist" in message.lower() or "no data" in message.lower():
        return "股票代號不存在"
    if "WinError 10013" in message:
        return "Windows socket 權限問題"
    if "connection failed" in message.lower():
        return "HTTPS 連線失敗"
    return "其他未知錯誤"


def fetch_twse_latest(config: dict[str, Any], timeout: int, logger: logging.Logger) -> pd.DataFrame:
    url = config["api"]["twse_stock_day_all"]
    data = request_json(url, None, timeout, logger)
    rows = []
    for item in data if isinstance(data, list) else []:
        code = norm_text(item.get("Code") or item.get("證券代號"))
        if not code:
            continue
        rows.append(
            {
                "stock_id": code.zfill(4),
                "stock_name": norm_text(item.get("Name") or item.get("證券名稱")),
                "open": parse_number(item.get("OpeningPrice") or item.get("開盤價")),
                "high": parse_number(item.get("HighestPrice") or item.get("最高價")),
                "low": parse_number(item.get("LowestPrice") or item.get("最低價")),
                "close": parse_number(item.get("ClosingPrice") or item.get("收盤價")),
                "volume": parse_number(item.get("TradeVolume") or item.get("成交股數")),
                "source": "TWSE OpenAPI STOCK_DAY_ALL",
                "api_url": url,
            }
        )
    return pd.DataFrame(rows)


def parse_number(value: Any) -> float | None:
    if value in (None, "", "--"):
        return None
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return None


def load_private_positions() -> dict[str, dict[str, Any]]:
    path = ROOT / "config" / "positions_private.json"
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    positions: dict[str, dict[str, Any]] = {}
    for item in data if isinstance(data, list) else []:
        code = norm_text(item.get("code"))
        if code:
            positions[code.zfill(4) if code.isdigit() else code] = item
    return positions


def fetch_finmind_history(
    stock_id: str,
    start_date: date,
    end_date: date,
    config: dict[str, Any],
    timeout: int,
    logger: logging.Logger,
) -> tuple[pd.DataFrame, str]:
    url = config["api"]["finmind_data"]
    params = {
        "dataset": "TaiwanStockPrice",
        "data_id": stock_id,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
    }
    data = request_json(url, params, timeout, logger)
    raw_path = ROOT / config["raw_data_folder"] / f"finmind_{stock_id}_{end_date.isoformat()}.json"
    raw_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    if not isinstance(data, dict) or data.get("status") not in (200, "200") or not data.get("data"):
        message = data.get("msg") if isinstance(data, dict) else "API 回傳格式不符"
        raise RuntimeError(f"FinMind 無有效資料：{message}")
    df = pd.DataFrame(data["data"])
    df = df.rename(
        columns={
            "date": "date",
            "stock_id": "stock_id",
            "Trading_Volume": "volume",
            "Trading_money": "value",
            "open": "open",
            "max": "high",
            "min": "low",
            "close": "close",
            "Trading_turnover": "turnover",
        }
    )
    keep = ["date", "stock_id", "open", "high", "low", "close", "volume", "value", "turnover"]
    for col in keep:
        if col not in df.columns:
            df[col] = None
    df = df[keep].copy()
    df["date"] = pd.to_datetime(df["date"])
    for col in ["open", "high", "low", "close", "volume", "value", "turnover"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna(subset=["date", "close"]).sort_values("date")
    return df, str(raw_path)


def fetch_finmind_dataset(
    dataset: str,
    stock_id: str,
    start_date: date,
    end_date: date,
    config: dict[str, Any],
    timeout: int,
    logger: logging.Logger,
    params_extra: dict[str, Any] | None = None,
) -> tuple[pd.DataFrame, str]:
    url = config["api"]["finmind_data"]
    params = {
        "dataset": dataset,
        "data_id": stock_id,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
    }
    if params_extra:
        params.update(params_extra)
    data = request_json(url, params, timeout, logger)
    raw_path = ROOT / config["raw_data_folder"] / f"finmind_{dataset}_{stock_id}_{end_date.isoformat()}.json"
    raw_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    if not isinstance(data, dict) or data.get("status") not in (200, "200") or not data.get("data"):
        message = data.get("msg") if isinstance(data, dict) else "API 回傳格式不符"
        raise RuntimeError(f"FinMind {dataset} 無有效資料：{message}")
    return pd.DataFrame(data["data"]), str(raw_path)


def add_indicators(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy().sort_values("date").reset_index(drop=True)
    for window in [5, 10, 20, 60, 120, 240]:
        df[f"MA{window}"] = df["close"].rolling(window).mean()
    df["vol_ma5"] = df["volume"].rolling(5).mean()
    df["vol_ma20"] = df["volume"].rolling(20).mean()

    low9 = df["low"].rolling(9).min()
    high9 = df["high"].rolling(9).max()
    rsv = (df["close"] - low9) / (high9 - low9) * 100
    k_values, d_values = [], []
    k_prev = 50.0
    d_prev = 50.0
    for value in rsv:
        if pd.isna(value):
            k_values.append(math.nan)
            d_values.append(math.nan)
            continue
        k_prev = k_prev * 2 / 3 + float(value) / 3
        d_prev = d_prev * 2 / 3 + k_prev / 3
        k_values.append(k_prev)
        d_values.append(d_prev)
    df["K"] = k_values
    df["D"] = d_values
    df["J"] = 3 * df["K"] - 2 * df["D"]

    ema12 = df["close"].ewm(span=12, adjust=False).mean()
    ema26 = df["close"].ewm(span=26, adjust=False).mean()
    df["MACD"] = ema12 - ema26
    df["MACD Signal"] = df["MACD"].ewm(span=9, adjust=False).mean()
    df["MACD Histogram"] = df["MACD"] - df["MACD Signal"]

    delta = df["close"].diff()
    gain = delta.clip(lower=0).rolling(14).mean()
    loss = (-delta.clip(upper=0)).rolling(14).mean()
    rs = gain / loss
    df["RSI14"] = 100 - (100 / (1 + rs))

    df["20日漲跌幅"] = df["close"] / df["close"].shift(20) - 1
    df["60日漲跌幅"] = df["close"] / df["close"].shift(60) - 1
    df["120日漲跌幅"] = df["close"] / df["close"].shift(120) - 1
    df["近60日高點回落幅度"] = df["close"] / df["high"].rolling(60).max() - 1
    df["近120日高點回落幅度"] = df["close"] / df["high"].rolling(120).max() - 1
    return df


def is_valid_number(value: Any) -> bool:
    return value is not None and not pd.isna(value)


def evaluate_status(latest: pd.Series, previous: pd.Series | None, row_count: int) -> tuple[str, list[str], str, dict[str, str]]:
    required = ["MA20", "MA60", "MA120", "K", "D", "MACD", "MACD Signal", "MACD Histogram", "vol_ma20", "近60日高點回落幅度", "近120日高點回落幅度"]
    if row_count < 120 or any(not is_valid_number(latest.get(col)) for col in required):
        return "資料不足", ["歷史資料不足，至少需 120 個交易日；MA240 完整判斷需 240 個交易日"], "是", {
            "跌破MA20": "資料不足",
            "跌破MA60": "資料不足",
            "跌破MA120": "資料不足",
            "放量下跌": "資料不足",
            "高檔爆量長黑": "資料不足",
        }

    flags = {
        "跌破MA20": "是" if latest["close"] < latest["MA20"] else "否",
        "跌破MA60": "是" if latest["close"] < latest["MA60"] else "否",
        "跌破MA120": "是" if latest["close"] < latest["MA120"] else "否",
        "放量下跌": "是" if latest["close"] < latest["open"] and latest["volume"] > latest["vol_ma20"] * 1.5 else "否",
        "高檔爆量長黑": "是" if latest["close"] < latest["open"] and latest["volume"] > latest["vol_ma20"] * 1.5 and latest["近60日高點回落幅度"] > -0.08 else "否",
    }

    status = "保留"
    reasons: list[str] = []

    def raise_status(new_status: str, reason: str) -> None:
        nonlocal status
        priority = {"保留": 0, "觀察": 1, "減碼警訊": 2, "出場警訊": 3, "資料不足": 4}
        if priority[new_status] > priority[status]:
            status = new_status
        reasons.append(reason)

    if flags["跌破MA20"] == "是":
        raise_status("觀察", "收盤價跌破 MA20")
    if flags["跌破MA60"] == "是":
        raise_status("減碼警訊", "收盤價跌破 MA60")
    if flags["跌破MA120"] == "是":
        raise_status("出場警訊", "收盤價跌破 MA120")

    if previous is not None and all(is_valid_number(previous.get(col)) for col in ["MACD", "MACD Signal", "MACD Histogram"]):
        macd_dead = previous["MACD"] >= previous["MACD Signal"] and latest["MACD"] < latest["MACD Signal"]
        hist_expanding_down = latest["MACD Histogram"] < previous["MACD Histogram"] < 0
        if macd_dead and hist_expanding_down:
            raise_status("減碼警訊", "MACD 死亡交叉且柱狀體連續擴大")

    if previous is not None and all(is_valid_number(previous.get(col)) and is_valid_number(latest.get(col)) for col in ["K", "D"]):
        if previous["K"] >= previous["D"] and latest["K"] < latest["D"] and latest["K"] > 80:
            raise_status("觀察", "KDJ 高檔死亡交叉")

    if flags["放量下跌"] == "是":
        raise_status("減碼警訊", "成交量大於 20 日均量 1.5 倍且收黑 K")
    if flags["高檔爆量長黑"] == "是":
        raise_status("減碼警訊", "高檔爆量長黑")

    if latest["近60日高點回落幅度"] <= -0.15:
        raise_status("減碼警訊", "從近 60 日高點回落超過 15%")
    elif latest["近60日高點回落幅度"] <= -0.10:
        raise_status("觀察", "從近 60 日高點回落超過 10%")

    if latest["近120日高點回落幅度"] <= -0.20:
        raise_status("出場警訊", "從近 120 日高點回落超過 20%")

    if not reasons:
        reasons.append("未觸發第一版技術/量價警訊")
    manual = "是" if status in {"減碼警訊", "出場警訊", "資料不足"} else "否"
    return status, reasons, manual, flags


def fmt(value: Any, digits: int = 2) -> Any:
    if value is None or pd.isna(value):
        return "資料不足"
    try:
        return round(float(value), digits)
    except (TypeError, ValueError):
        return value


def pct(value: Any) -> Any:
    if value is None or pd.isna(value):
        return "資料不足"
    return round(float(value), 4)


def status_level(status: str) -> int:
    return {"保留": 0, "保留但觀察": 1, "觀察": 1, "減碼警訊": 2, "出場警訊": 3, "資料不足": 4}.get(status, 0)


def worse_status(*statuses: str) -> str:
    order = ["保留", "觀察", "減碼警訊", "出場警訊", "資料不足"]
    best = "保留"
    for status in statuses:
        if status_level(status) > status_level(best):
            best = "觀察" if status == "保留但觀察" else status
    return best


def safe_ratio(current: Any, previous: Any) -> float | None:
    if not is_valid_number(current) or not is_valid_number(previous) or float(previous) == 0:
        return None
    return float(current) / float(previous) - 1


def count_recent_signs(values: list[float], positive: bool) -> int:
    count = 0
    for value in reversed(values):
        if pd.isna(value):
            break
        if positive and value > 0:
            count += 1
        elif not positive and value < 0:
            count += 1
        else:
            break
    return count


def build_monthly_revenue_row(
    holding: Holding,
    df: pd.DataFrame | None,
    technical_row: dict[str, Any],
    source: str,
    api_status: str,
    error: str = "",
) -> dict[str, Any]:
    base = {
        "公司代號": holding.stock_id,
        "公司名稱": holding.stock_name,
        "資料年月": "資料不足",
        "當月營收": "資料不足",
        "上月營收": "資料不足",
        "去年同月營收": "資料不足",
        "月增率": "資料不足",
        "年增率": "資料不足",
        "累計營收": "資料不足",
        "去年累計營收": "資料不足",
        "累計年增率": "資料不足",
        "連續幾個月年增": "資料不足",
        "連續幾個月年減": "資料不足",
        "基本面警訊": "資料不足",
        "基本面狀態": "資料不足",
        "資料來源": source or "資料不足",
        "API更新狀態": api_status,
    }
    if df is None or df.empty:
        base["基本面警訊"] = error or "月營收資料不足"
        return base

    work = df.copy()
    for col in ["revenue", "revenue_year", "revenue_month"]:
        if col not in work.columns:
            base["基本面警訊"] = f"月營收欄位不足：缺 {col}"
            return base
    work["date"] = pd.to_datetime(work["date"], errors="coerce")
    work["revenue"] = pd.to_numeric(work["revenue"], errors="coerce")
    work = work.dropna(subset=["date", "revenue"]).sort_values("date")
    if work.empty:
        base["基本面警訊"] = "月營收資料不足"
        return base

    work["ym"] = work["date"].dt.strftime("%Y-%m")
    work["month"] = work["date"].dt.month
    work["year"] = work["date"].dt.year
    work["prev_revenue"] = work["revenue"].shift(1)
    work["last_year_revenue"] = work.groupby("month")["revenue"].shift(1)
    work["mom"] = work.apply(lambda r: safe_ratio(r["revenue"], r["prev_revenue"]), axis=1)
    work["yoy"] = work.apply(lambda r: safe_ratio(r["revenue"], r["last_year_revenue"]), axis=1)
    work["cum_revenue"] = work.groupby("year")["revenue"].cumsum()
    last_year_cum = work[["year", "month", "cum_revenue"]].copy()
    last_year_cum["year"] = last_year_cum["year"] + 1
    last_year_cum = last_year_cum.rename(columns={"cum_revenue": "last_year_cum_revenue"})
    work = work.merge(last_year_cum, on=["year", "month"], how="left")
    work["cum_yoy"] = work.apply(lambda r: safe_ratio(r["cum_revenue"], r["last_year_cum_revenue"]), axis=1)

    latest = work.iloc[-1]
    previous = work.iloc[-2] if len(work) >= 2 else None
    yoy_values = work["yoy"].tolist()
    up_count = count_recent_signs(yoy_values, True)
    down_count = count_recent_signs(yoy_values, False)
    reasons: list[str] = []
    status = "保留"
    if down_count >= 3:
        status = worse_status(status, "減碼警訊")
        reasons.append("月營收連續 3 個月年減")
    elif down_count >= 2:
        status = worse_status(status, "觀察")
        reasons.append("月營收連續 2 個月年減")
    if previous is not None and is_valid_number(previous.get("cum_yoy")) and is_valid_number(latest.get("cum_yoy")):
        if previous["cum_yoy"] > 0 and latest["cum_yoy"] < 0:
            status = worse_status(status, "觀察")
            reasons.append("累計營收年增率由正轉負")
    tech_weak = technical_row.get("跌破MA60") == "是"
    if previous is not None and is_valid_number(previous.get("yoy")) and is_valid_number(latest.get("yoy")):
        if previous["yoy"] > 0 and latest["yoy"] < 0 and tech_weak:
            status = worse_status(status, "減碼警訊")
            reasons.append("月營收年增率由正轉負且股價跌破 MA60")
    if not reasons:
        reasons.append("未觸發月營收警訊")

    base.update(
        {
            "資料年月": latest["ym"],
            "當月營收": fmt(latest["revenue"], 0),
            "上月營收": fmt(latest.get("prev_revenue"), 0),
            "去年同月營收": fmt(latest.get("last_year_revenue"), 0),
            "月增率": pct(latest.get("mom")),
            "年增率": pct(latest.get("yoy")),
            "累計營收": fmt(latest.get("cum_revenue"), 0),
            "去年累計營收": fmt(latest.get("last_year_cum_revenue"), 0),
            "累計年增率": pct(latest.get("cum_yoy")),
            "連續幾個月年增": up_count,
            "連續幾個月年減": down_count,
            "基本面警訊": "；".join(reasons),
            "基本面狀態": status,
        }
    )
    return base


def build_institutional_row(
    holding: Holding,
    df: pd.DataFrame | None,
    technical_row: dict[str, Any],
    source: str,
    api_status: str,
    error: str = "",
) -> dict[str, Any]:
    base = {
        "日期": "資料不足",
        "公司代號": holding.stock_id,
        "公司名稱": holding.stock_name,
        "外資買賣超": "資料不足",
        "投信買賣超": "資料不足",
        "自營商買賣超": "資料不足",
        "三大法人合計買賣超": "資料不足",
        "近5日外資買賣超": "資料不足",
        "近5日投信買賣超": "資料不足",
        "近20日三大法人買賣超": "資料不足",
        "是否連續賣超": "資料不足",
        "籌碼警訊": error or "法人籌碼資料不足",
        "籌碼狀態": "資料不足",
        "資料來源": source or "資料不足",
        "API更新狀態": api_status,
    }
    if df is None or df.empty:
        return base

    work = df.copy()
    if "date" not in work.columns:
        return base
    work["date"] = pd.to_datetime(work["date"], errors="coerce")
    if "buy" in work.columns and "sell" in work.columns:
        work["net"] = pd.to_numeric(work["buy"], errors="coerce") - pd.to_numeric(work["sell"], errors="coerce")
    elif "buy_sell" in work.columns:
        work["net"] = pd.to_numeric(work["buy_sell"], errors="coerce")
    elif "buy_sell_amount" in work.columns:
        work["net"] = pd.to_numeric(work["buy_sell_amount"], errors="coerce")
    else:
        return base
    name_col = "name" if "name" in work.columns else ("institutional_investors" if "institutional_investors" in work.columns else None)
    if not name_col:
        return base

    def category(name: Any) -> str | None:
        text = norm_text(name)
        if "外資" in text or "Foreign" in text:
            return "foreign"
        if "投信" in text or "Investment" in text:
            return "trust"
        if "自營" in text or "Dealer" in text:
            return "dealer"
        return None

    work["category"] = work[name_col].map(category)
    work = work.dropna(subset=["date", "category"]).sort_values("date")
    if work.empty:
        return base
    pivot = work.pivot_table(index="date", columns="category", values="net", aggfunc="sum").fillna(0)
    for col in ["foreign", "trust", "dealer"]:
        if col not in pivot.columns:
            pivot[col] = 0
    pivot["total"] = pivot["foreign"] + pivot["trust"] + pivot["dealer"]
    pivot = pivot.sort_index()
    latest = pivot.iloc[-1]
    last5 = pivot.tail(5)
    last20 = pivot.tail(20)
    foreign_5 = float(last5["foreign"].sum())
    trust_5 = float(last5["trust"].sum())
    total_20 = float(last20["total"].sum())
    foreign_5_days_sell = len(last5) >= 5 and bool((last5["foreign"] < 0).all())
    trust_5_days_sell = len(last5) >= 5 and bool((last5["trust"] < 0).all())
    total_20_sell = total_20 < 0 and abs(total_20) > max(abs(float(last20["total"].mean())) * 10, 1)
    tech_ma60 = technical_row.get("跌破MA60") == "是"
    tech_ma120 = technical_row.get("跌破MA120") == "是"

    reasons: list[str] = []
    status = "保留"
    if foreign_5_days_sell:
        status = worse_status(status, "觀察")
        reasons.append("外資連續 5 日賣超")
    if foreign_5_days_sell and trust_5_days_sell:
        status = worse_status(status, "減碼警訊")
        reasons.append("外資與投信同時連續賣超")
    if total_20_sell:
        status = worse_status(status, "減碼警訊")
        reasons.append("近 20 日三大法人明顯賣超")
    if tech_ma60 and latest["total"] < 0:
        status = worse_status(status, "減碼警訊")
        reasons.append("股價跌破 MA60 且法人同步賣超")
    if tech_ma120 and foreign_5_days_sell:
        status = worse_status(status, "出場警訊")
        reasons.append("股價跌破 MA120 且法人連續賣超")
    if not reasons:
        reasons.append("未觸發法人籌碼警訊")

    base.update(
        {
            "日期": pivot.index[-1].date().isoformat(),
            "外資買賣超": fmt(latest["foreign"], 0),
            "投信買賣超": fmt(latest["trust"], 0),
            "自營商買賣超": fmt(latest["dealer"], 0),
            "三大法人合計買賣超": fmt(latest["total"], 0),
            "近5日外資買賣超": fmt(foreign_5, 0),
            "近5日投信買賣超": fmt(trust_5, 0),
            "近20日三大法人買賣超": fmt(total_20, 0),
            "是否連續賣超": "是" if foreign_5_days_sell or trust_5_days_sell else "否",
            "籌碼警訊": "；".join(reasons),
            "籌碼狀態": status,
        }
    )
    return base


def extract_financial_metric(df: pd.DataFrame, keywords: list[str]) -> float | None:
    if df is None or df.empty:
        return None
    type_col = "type" if "type" in df.columns else ("origin_name" if "origin_name" in df.columns else None)
    value_col = "value" if "value" in df.columns else None
    if not type_col or not value_col:
        return None
    mask = pd.Series(False, index=df.index)
    names = df[type_col].astype(str)
    for keyword in keywords:
        mask = mask | names.str.contains(keyword, case=False, na=False)
    values = pd.to_numeric(df.loc[mask, value_col], errors="coerce").dropna()
    if values.empty:
        return None
    return float(values.iloc[-1])


def build_financial_row(
    holding: Holding,
    financial_df: pd.DataFrame | None,
    per_df: pd.DataFrame | None,
    technical_row: dict[str, Any],
    source: str,
    api_status: str,
    error: str = "",
) -> dict[str, Any]:
    base = {
        "公司代號": holding.stock_id,
        "公司名稱": holding.stock_name,
        "EPS": "資料不足",
        "近四季EPS": "資料不足",
        "毛利率": "資料不足",
        "營業利益率": "資料不足",
        "淨利率": "資料不足",
        "本益比": "資料不足",
        "股價淨值比": "資料不足",
        "殖利率": "資料不足",
        "財報更新日期": "資料不足",
        "財報警訊": error or "財報與估值資料不足",
        "財報狀態": "資料不足",
        "公司下修展望": "資料不足",
        "資料來源": source or "資料不足",
        "API更新狀態": api_status,
    }
    if financial_df is None or financial_df.empty:
        return base

    work = financial_df.copy()
    if "date" in work.columns:
        work["date"] = pd.to_datetime(work["date"], errors="coerce")
        work = work.sort_values("date")
    type_col = "type" if "type" in work.columns else ("origin_name" if "origin_name" in work.columns else None)
    value_col = "value" if "value" in work.columns else None
    eps_series = pd.Series(dtype=float)
    if type_col and value_col:
        eps_mask = work[type_col].astype(str).str.contains("EPS|每股盈餘|基本每股盈餘", case=False, na=False)
        eps_series = pd.to_numeric(work.loc[eps_mask, value_col], errors="coerce").dropna().tail(4)
    eps = float(eps_series.iloc[-1]) if not eps_series.empty else None
    eps_text = " / ".join(f"{v:.2f}" for v in eps_series.tolist()) if not eps_series.empty else "資料不足"

    gross_margin = extract_financial_metric(work, ["GrossProfitMargin", "毛利率"])
    operating_margin = extract_financial_metric(work, ["OperatingIncomeRatio", "營業利益率", "營益率"])
    net_margin = extract_financial_metric(work, ["NetIncomeRatio", "淨利率"])

    per_latest = None
    if per_df is not None and not per_df.empty:
        per_work = per_df.copy()
        if "date" in per_work.columns:
            per_work["date"] = pd.to_datetime(per_work["date"], errors="coerce")
            per_work = per_work.sort_values("date")
        per_latest = per_work.iloc[-1]

    pe = parse_number(per_latest.get("PER")) if per_latest is not None and "PER" in per_latest else None
    pbr = parse_number(per_latest.get("PBR")) if per_latest is not None and "PBR" in per_latest else None
    dy = parse_number(per_latest.get("dividend_yield")) if per_latest is not None and "dividend_yield" in per_latest else None
    if dy is None and per_latest is not None and "殖利率" in per_latest:
        dy = parse_number(per_latest.get("殖利率"))

    reasons: list[str] = []
    status = "保留"
    if len(eps_series) >= 3 and eps_series.iloc[-1] < eps_series.iloc[-2] < eps_series.iloc[-3]:
        status = worse_status(status, "減碼警訊")
        reasons.append("EPS 連續 2 季衰退")
    if gross_margin is None:
        reasons.append("毛利率資料不足")
    if pe is not None and pe > 40 and technical_row.get("系統狀態") in {"觀察", "減碼警訊", "出場警訊"}:
        status = worse_status(status, "觀察")
        reasons.append("本益比偏高且技術面轉弱")
    # 公司下修展望需由新聞/法說摘要或人工欄位判讀，第一版先保留資料不足。
    if not reasons:
        reasons.append("未觸發財報估值警訊")

    base.update(
        {
            "EPS": fmt(eps),
            "近四季EPS": eps_text,
            "毛利率": pct(gross_margin / 100 if gross_margin and abs(gross_margin) > 1 else gross_margin),
            "營業利益率": pct(operating_margin / 100 if operating_margin and abs(operating_margin) > 1 else operating_margin),
            "淨利率": pct(net_margin / 100 if net_margin and abs(net_margin) > 1 else net_margin),
            "本益比": fmt(pe),
            "股價淨值比": fmt(pbr),
            "殖利率": pct(dy / 100 if dy and abs(dy) > 1 else dy),
            "財報更新日期": work["date"].max().date().isoformat() if "date" in work.columns and pd.notna(work["date"].max()) else "資料不足",
            "財報警訊": "；".join(reasons),
            "財報狀態": status,
            "公司下修展望": "資料不足",
        }
    )
    return base


def build_news_row(
    holding: Holding,
    df: pd.DataFrame | None,
    source: str,
    api_status: str,
    error: str = "",
) -> dict[str, Any]:
    base = {
        "公司代號": holding.stock_id,
        "公司名稱": holding.stock_name,
        "公司新聞標題": "資料不足",
        "新聞日期": "資料不足",
        "新聞來源": "資料不足",
        "新聞連結": "資料不足",
        "新聞摘要": error or "新聞資料不足；可於本表人工補充法說會、產業與重大新聞",
        "正面/中性/負面": "資料不足",
        "是否重大利空": "資料不足",
        "產業趨勢摘要": "資料不足",
        "新聞風險警訊": error or "新聞資料不足",
        "新聞風險狀態": "資料不足",
        "資料來源": source or "資料不足",
        "API更新狀態": api_status,
    }
    if df is None or df.empty:
        return base

    work = df.copy()
    date_col = "date" if "date" in work.columns else ("publish_date" if "publish_date" in work.columns else None)
    title_col = "title" if "title" in work.columns else ("新聞標題" if "新聞標題" in work.columns else None)
    link_col = "link" if "link" in work.columns else ("url" if "url" in work.columns else None)
    source_col = "source" if "source" in work.columns else ("publisher" if "publisher" in work.columns else None)
    summary_col = "summary" if "summary" in work.columns else ("content" if "content" in work.columns else None)
    if not title_col:
        return base
    if date_col:
        work[date_col] = pd.to_datetime(work[date_col], errors="coerce")
        work = work.sort_values(date_col)
    latest = work.iloc[-1]
    text_blob = " ".join(norm_text(latest.get(col)) for col in [title_col, summary_col] if col)
    negative_keywords = ["下修", "衰退", "虧損", "違約", "停工", "裁員", "制裁", "調查", "重大訊息", "利空", "需求轉弱", "競爭加劇"]
    positive_keywords = ["上修", "成長", "創高", "訂單", "擴產", "合作", "得標", "獲利"]
    is_negative = any(k in text_blob for k in negative_keywords)
    is_positive = any(k in text_blob for k in positive_keywords)
    sentiment = "負面" if is_negative else ("正面" if is_positive else "中性")
    status = "觀察" if is_negative else "保留"
    reasons = ["重大利空新聞或負面關鍵字"] if is_negative else ["未偵測重大利空新聞"]
    if "下修" in text_blob or "需求轉弱" in text_blob:
        status = "減碼警訊"
        reasons.append("法說會下修展望或產業需求轉弱")
    if "競爭優勢" in text_blob and ("改變" in text_blob or "喪失" in text_blob):
        status = "出場警訊"
        reasons.append("公司競爭優勢改變")

    base.update(
        {
            "公司新聞標題": norm_text(latest.get(title_col)),
            "新聞日期": latest.get(date_col).date().isoformat() if date_col and hasattr(latest.get(date_col), "date") else "資料不足",
            "新聞來源": norm_text(latest.get(source_col)) if source_col else "資料不足",
            "新聞連結": norm_text(latest.get(link_col)) if link_col else "資料不足",
            "新聞摘要": norm_text(latest.get(summary_col))[:300] if summary_col else norm_text(latest.get(title_col)),
            "正面/中性/負面": sentiment,
            "是否重大利空": "是" if is_negative else "否",
            "產業趨勢摘要": "需人工補充產業趨勢；第三階段可接新聞摘要 API",
            "新聞風險警訊": "；".join(reasons),
            "新聞風險狀態": status,
        }
    )
    return base


def score_from_status(status: str, max_points: int) -> int:
    if status == "保留":
        return max_points
    if status in {"保留但觀察", "觀察"}:
        return round(max_points * 0.7)
    if status == "減碼警訊":
        return round(max_points * 0.4)
    if status == "出場警訊":
        return round(max_points * 0.15)
    return round(max_points * 0.3)


def build_score_row(
    holding: Holding,
    tech: dict[str, Any],
    revenue: dict[str, Any],
    chip: dict[str, Any],
    financial: dict[str, Any],
    news: dict[str, Any],
) -> dict[str, Any]:
    tech_score = score_from_status(tech.get("系統狀態", "資料不足"), 30)
    fundamental_status = worse_status(revenue.get("基本面狀態", "資料不足"), financial.get("財報狀態", "資料不足"))
    fundamental_score = score_from_status(fundamental_status, 30)
    chip_score = score_from_status(chip.get("籌碼狀態", "資料不足"), 20)
    news_score = score_from_status(news.get("新聞風險狀態", "資料不足"), 20)
    total = tech_score + fundamental_score + chip_score + news_score
    if total >= 80:
        status = "保留"
    elif total >= 65:
        status = "保留但觀察"
    elif total >= 50:
        status = "觀察"
    elif total >= 40:
        status = "減碼警訊"
    else:
        status = "出場警訊"
    severe = worse_status(tech.get("系統狀態", "保留"), fundamental_status, chip.get("籌碼狀態", "保留"), news.get("新聞風險狀態", "保留"))
    if severe in {"減碼警訊", "出場警訊"} and status_level(severe) > status_level(status):
        status = severe
    manual = "是" if status in {"減碼警訊", "出場警訊"} or "資料不足" in [revenue.get("基本面狀態"), chip.get("籌碼狀態"), financial.get("財報狀態"), news.get("新聞風險狀態")] else "否"
    reasons = [
        f"技術：{tech.get('主要理由', '資料不足')}",
        f"月營收：{revenue.get('基本面警訊', '資料不足')}",
        f"財報估值：{financial.get('財報警訊', '資料不足')}",
        f"籌碼：{chip.get('籌碼警訊', '資料不足')}",
        f"新聞：{news.get('新聞風險警訊', '資料不足')}",
    ]
    return {
        "公司代號": holding.stock_id,
        "公司名稱": holding.stock_name,
        "技術面分數": tech_score,
        "基本面分數": fundamental_score,
        "籌碼面分數": chip_score,
        "新聞與產業分數": news_score,
        "總分": total,
        "第二階段狀態": status,
        "主要理由": "；".join(reasons),
        "需要人工確認": manual,
        "不可自動下單": "是",
        "不可直接執行交易": "是",
        "資料來源與日期": f"技術{tech.get('更新日期','資料不足')}；營收{revenue.get('資料年月','資料不足')}；法人{chip.get('日期','資料不足')}；財報{financial.get('財報更新日期','資料不足')}；新聞{news.get('新聞日期','資料不足')}",
    }


def build_result_row(holding: Holding, df: pd.DataFrame | None, source: str, api_status: str, error: str = "") -> dict[str, Any]:
    if df is None or df.empty:
        latest_date = date.today().isoformat()
        return {
            "股票代號": holding.stock_id,
            "股票名稱": holding.stock_name,
            "市場": holding.market,
            "投資分類": holding.category,
            "持有股數": holding.shares or 0,
            "成本均價": holding.average_cost or 0,
            "最新價格": "資料不足",
            "市值": "資料不足",
            "未實現損益": "資料不足",
            "未實現損益率": "資料不足",
            "更新日期": latest_date,
            "開盤價": "資料不足",
            "最高價": "資料不足",
            "最低價": "資料不足",
            "收盤價": "資料不足",
            "成交量": "資料不足",
            "MA5": "資料不足",
            "MA10": "資料不足",
            "MA20": "資料不足",
            "MA60": "資料不足",
            "MA120": "資料不足",
            "MA240": "資料不足",
            "K": "資料不足",
            "D": "資料不足",
            "J": "資料不足",
            "MACD": "資料不足",
            "MACD Signal": "資料不足",
            "MACD Histogram": "資料不足",
            "RSI14": "資料不足",
            "20日漲跌幅": "資料不足",
            "60日漲跌幅": "資料不足",
            "120日漲跌幅": "資料不足",
            "跌破MA20": "資料不足",
            "跌破MA60": "資料不足",
            "跌破MA120": "資料不足",
            "放量下跌": "資料不足",
            "高檔爆量長黑": "資料不足",
            "系統狀態": "資料不足",
            "主要理由": error or "API 失敗或歷史資料不足",
            "是否需要人工確認": "是",
            "資料來源": source or "資料不足",
            "API更新狀態": api_status,
        }

    latest = df.iloc[-1]
    previous = df.iloc[-2] if len(df) >= 2 else None
    status, reasons, manual, flags = evaluate_status(latest, previous, len(df))
    close = latest["close"]
    shares = holding.shares or 0
    cost = holding.average_cost or 0
    market_value = close * shares if shares and is_valid_number(close) else None
    unrealized = (close - cost) * shares if shares and cost and is_valid_number(close) else None
    unrealized_rate = (close / cost - 1) if cost and is_valid_number(close) else None

    return {
        "股票代號": holding.stock_id,
        "股票名稱": holding.stock_name,
        "市場": holding.market,
        "投資分類": holding.category,
        "持有股數": shares,
        "成本均價": cost,
        "最新價格": fmt(close),
        "市值": fmt(market_value),
        "未實現損益": fmt(unrealized),
        "未實現損益率": pct(unrealized_rate),
        "更新日期": latest["date"].date().isoformat() if hasattr(latest["date"], "date") else str(latest["date"]),
        "開盤價": fmt(latest["open"]),
        "最高價": fmt(latest["high"]),
        "最低價": fmt(latest["low"]),
        "收盤價": fmt(close),
        "成交量": fmt(latest["volume"], 0),
        "MA5": fmt(latest["MA5"]),
        "MA10": fmt(latest["MA10"]),
        "MA20": fmt(latest["MA20"]),
        "MA60": fmt(latest["MA60"]),
        "MA120": fmt(latest["MA120"]),
        "MA240": fmt(latest["MA240"]),
        "K": fmt(latest["K"]),
        "D": fmt(latest["D"]),
        "J": fmt(latest["J"]),
        "MACD": fmt(latest["MACD"], 4),
        "MACD Signal": fmt(latest["MACD Signal"], 4),
        "MACD Histogram": fmt(latest["MACD Histogram"], 4),
        "RSI14": fmt(latest["RSI14"]),
        "20日漲跌幅": pct(latest["20日漲跌幅"]),
        "60日漲跌幅": pct(latest["60日漲跌幅"]),
        "120日漲跌幅": pct(latest["120日漲跌幅"]),
        "跌破MA20": flags["跌破MA20"],
        "跌破MA60": flags["跌破MA60"],
        "跌破MA120": flags["跌破MA120"],
        "放量下跌": flags["放量下跌"],
        "高檔爆量長黑": flags["高檔爆量長黑"],
        "系統狀態": status,
        "主要理由": "；".join(reasons),
        "是否需要人工確認": manual,
        "資料來源": source,
        "API更新狀態": api_status,
    }


def clear_and_write(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.cell(1, 1, "自動更新結果：由 scripts/update_stock_data.py 產生；手動持股欄位不在此修改")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    for c, h in enumerate(headers, 1):
        ws.cell(2, c, h)
    for r, row in enumerate(rows, 3):
        for c, h in enumerate(headers, 1):
            ws.cell(r, c, row.get(h, ""))
    style_table(ws, 2, 3, max(3, len(rows) + 2), len(headers))


def style_table(ws, header_row: int, body_start: int, body_end: int, max_col: int) -> None:
    blue_fill = PatternFill("solid", fgColor="2F6F9F")
    gray_fill = PatternFill("solid", fgColor="F3F6F8")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="16324F")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for col in range(1, max_col + 1):
        cell = ws.cell(header_row, col)
        cell.fill = blue_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 16
    for row in range(body_start, body_end + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row % 2 == 0:
                cell.fill = gray_fill
    ws.freeze_panes = "A3"


def write_source_sheet(ws, run_date: str, rows: list[dict[str, Any]]) -> None:
    headers = ["日期", "股票代號", "資料類型", "來源名稱", "來源連結/檔名", "資料期間", "重點摘要", "備註"]
    ws.delete_rows(1, ws.max_row)
    ws.cell(1, 1, "資料來源與更新紀錄")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    for c, h in enumerate(headers, 1):
        ws.cell(2, c, h)
    for r, row in enumerate(rows, 3):
        ws.cell(r, 1, run_date)
        ws.cell(r, 2, row["股票代號"])
        ws.cell(r, 3, "日線/技術指標")
        ws.cell(r, 4, row["資料來源"])
        ws.cell(r, 5, row["API更新狀態"])
        ws.cell(r, 6, row["更新日期"])
        ws.cell(r, 7, row["主要理由"])
        ws.cell(r, 8, "減碼/出場/資料不足皆需人工確認")
    style_table(ws, 2, 3, max(3, len(rows) + 2), len(headers))


def update_manual_api_columns(wb: Any, config: dict[str, Any], rows_by_id: dict[str, dict[str, Any]]) -> None:
    ws = wb[config["manual_sheet"]]
    headers = {norm_text(ws.cell(2, col).value): col for col in range(1, ws.max_column + 1)}
    target_map = {
        "最新價格(API)": "最新價格",
        "更新日期(API)": "更新日期",
        "系統狀態(API/公式)": "系統狀態",
        "需要人工確認": "是否需要人工確認",
        "最後檢查日": "更新日期",
        "資料缺失提醒": "主要理由",
    }
    stock_col = headers.get("股票代號")
    tracking_col = headers.get("是否追蹤(Y/N)")
    if not stock_col:
        return
    for row in range(3, ws.max_row + 1):
        stock_id = norm_text(ws.cell(row, stock_col).value)
        if not stock_id:
            continue
        stock_id = stock_id.zfill(4) if stock_id.isdigit() else stock_id
        result = rows_by_id.get(stock_id)
        if not result:
            continue
        if tracking_col and norm_text(ws.cell(row, tracking_col).value).upper() != "Y":
            continue
        for excel_header, result_key in target_map.items():
            col = headers.get(excel_header)
            if col:
                ws.cell(row, col, result[result_key])


def write_report(rows: list[dict[str, Any]], report_path: Path, run_date: str) -> None:
    total_mv = sum(float(r["市值"]) for r in rows if isinstance(r.get("市值"), (int, float)))
    total_pl = sum(float(r["未實現損益"]) for r in rows if isinstance(r.get("未實現損益"), (int, float)))

    def section(status: str) -> list[dict[str, Any]]:
        return [r for r in rows if r["系統狀態"] == status]

    def lines(items: list[dict[str, Any]]) -> str:
        if not items:
            return "- 無\n"
        return "".join(f"- {r['股票代號']} {r['股票名稱']}：{r['主要理由']}\n" for r in items)

    failures = [r for r in rows if "成功" not in str(r["API更新狀態"])]
    manual = [r for r in rows if r["是否需要人工確認"] == "是"]
    content = [
        f"# 每日股票追蹤報告 {run_date}",
        "",
        "## 1. 今日更新日期與資料來源",
        f"- 更新日期：{run_date}",
        "- 資料來源：TWSE OpenAPI 最新行情架構、FinMind TaiwanStockPrice 歷史日線補充。實際來源逐檔記錄於 Excel「09_資料來源」與 log。",
        "",
        "## 2. 投資組合總市值",
        f"- 總市值：{total_mv:,.0f}",
        "",
        "## 3. 整體未實現損益",
        f"- 未實現損益：{total_pl:,.0f}",
        "",
        "## 4. 保留清單",
        lines(section("保留")),
        "## 5. 觀察清單",
        lines(section("觀察")),
        "## 6. 減碼警訊清單",
        lines(section("減碼警訊")),
        "## 7. 出場警訊清單",
        lines(section("出場警訊")),
        "## 8. 資料不足清單",
        lines(section("資料不足")),
        "## 9. 今日最需要人工確認的股票",
        lines(manual),
        "## 10. 每檔股票的主要理由",
        lines(rows),
        "## 11. API 失敗或資料缺失說明",
        lines(failures),
        "",
        "> 本報告為投資研究與風控提醒工具，不是自動交易工具；減碼與出場警訊都需要人工確認。",
    ]
    report_path.write_text("\n".join(content), encoding="utf-8")


def write_stage2_report(
    score_rows: list[dict[str, Any]],
    revenue_rows: list[dict[str, Any]],
    chip_rows: list[dict[str, Any]],
    news_rows: list[dict[str, Any]],
    report_path: Path,
    run_date: str,
) -> None:
    def by_status(statuses: set[str]) -> list[dict[str, Any]]:
        return [r for r in score_rows if r["第二階段狀態"] in statuses]

    def lines(items: list[dict[str, Any]], reason_key: str = "主要理由") -> str:
        if not items:
            return "- 無\n"
        return "".join(f"- {r.get('公司代號')} {r.get('公司名稱')}：{r.get(reason_key, '資料不足')}\n" for r in items)

    high_risk = by_status({"減碼警訊", "出場警訊"})
    weak_fundamental = [r for r in revenue_rows if r.get("基本面狀態") in {"觀察", "減碼警訊", "出場警訊"}]
    chip_sell = [r for r in chip_rows if r.get("籌碼狀態") in {"觀察", "減碼警訊", "出場警訊"}]
    bad_news = [r for r in news_rows if r.get("新聞風險狀態") in {"觀察", "減碼警訊", "出場警訊"}]
    keep = by_status({"保留", "保留但觀察"})

    content = [
        f"# 第二階段股票追蹤摘要 {run_date}",
        "",
        "## 更新重點",
        "- 已納入技術面、基本面、籌碼法人、新聞與產業風險四大構面。",
        "- 綜合評分為 100 分：技術面 30、基本面 30、籌碼面 20、新聞與產業 20。",
        "- 減碼警訊與出場警訊都只代表風控提醒，必須人工確認，不可自動下單。",
        "",
        "## 高風險股票清單",
        lines(high_risk),
        "## 基本面轉弱股票清單",
        lines(weak_fundamental, "基本面警訊"),
        "## 法人連續賣超股票清單",
        lines(chip_sell, "籌碼警訊"),
        "## 新聞利空股票清單",
        lines(bad_news, "新聞風險警訊"),
        "## 仍可保留股票清單",
        lines(keep),
        "## 每檔綜合評分",
    ]
    for row in score_rows:
        content.append(
            f"- {row['公司代號']} {row['公司名稱']}：總分 {row['總分']}，狀態 {row['第二階段狀態']}，人工確認 {row['需要人工確認']}。"
        )
    content.extend(
        [
            "",
            "## 資料不足提醒",
            "- 新聞與法說會展望若公開 API 無法取得，系統會標示資料不足，可在「14_新聞風險追蹤」人工補充。",
            "- 公司下修展望、競爭優勢改變等事件仍需人工確認公開資訊或法說會資料。",
            "",
            "> 本系統不登入券商、不自動下單、不保存帳號密碼或 OTP，不提供絕對買賣指令。",
        ]
    )
    report_path.write_text("\n".join(content), encoding="utf-8")


def split_reason_text(value: Any) -> list[str]:
    text = norm_text(value)
    if not text:
        return []
    parts: list[str] = []
    for chunk in text.replace("\n", "；").split("；"):
        item = chunk.strip(" -\t")
        if item:
            parts.append(item)
    return parts


def clean_json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float, str, bool)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def public_source(value: Any, publish_raw_paths: bool = False) -> str | None:
    text = norm_text(value)
    if not text:
        return None
    if publish_raw_paths:
        return text
    parts = []
    for chunk in text.split("；"):
        clean = chunk.strip()
        if not clean:
            continue
        if "raw=" in clean:
            clean = clean.split("；raw=", 1)[0].split("raw=", 1)[0].strip("； ")
        if "C:\\" in clean or "C:/" in clean:
            continue
        if clean:
            parts.append(clean)
    return "；".join(parts) or "FinMind / 本機更新腳本"


def public_detail_record(record: dict[str, Any], publish_raw_paths: bool = False) -> dict[str, Any]:
    hidden_keys = {
        "持有股數",
        "成本均價",
        "市值",
        "未實現損益",
        "資料來源",
        "API更新狀態",
    }
    public: dict[str, Any] = {}
    for key, value in record.items():
        if key in hidden_keys:
            continue
        if "來源" in key:
            public[key] = public_source(value, publish_raw_paths)
        else:
            public[key] = clean_json_value(value)
    return public


def pick_public(record: dict[str, Any], keys: list[str], publish_raw_paths: bool = False) -> dict[str, Any]:
    picked: dict[str, Any] = {}
    for key in keys:
        if key not in record:
            continue
        value = record.get(key)
        if "來源" in key:
            picked[key] = public_source(value, publish_raw_paths)
        else:
            picked[key] = clean_json_value(value)
    return picked


def export_dashboard_json(
    result_rows: list[dict[str, Any]],
    revenue_rows: list[dict[str, Any]],
    chip_rows: list[dict[str, Any]],
    financial_rows: list[dict[str, Any]],
    news_rows: list[dict[str, Any]],
    score_rows: list[dict[str, Any]],
    run_date: str,
    public_options: dict[str, Any] | None = None,
) -> Path:
    docs_data = ROOT / "docs" / "data"
    docs_data.mkdir(parents=True, exist_ok=True)
    public_options = public_options or {}
    view_settings = load_public_view_settings()
    publish_position_amounts = bool(view_settings.get("showTotalAssets", False))
    publish_shares = bool(view_settings.get("showShares", False))
    publish_average_cost = bool(view_settings.get("showCostBasis", False))
    publish_raw_paths = bool(public_options.get("publish_raw_source_paths", False))
    show_pnl_rate = bool(view_settings.get("showPnlRate", True))
    private_positions = load_private_positions()

    by_code = {str(row.get("股票代號") or row.get("公司代號")): row for row in result_rows}
    revenue_by_code = {str(row.get("公司代號")): row for row in revenue_rows}
    chip_by_code = {str(row.get("公司代號")): row for row in chip_rows}
    financial_by_code = {str(row.get("公司代號")): row for row in financial_rows}
    news_by_code = {str(row.get("公司代號")): row for row in news_rows}
    score_by_code = {str(row.get("公司代號")): row for row in score_rows}

    stocks: list[dict[str, Any]] = []
    for code, tech in by_code.items():
        score = score_by_code.get(code, {})
        revenue = revenue_by_code.get(code, {})
        chip = chip_by_code.get(code, {})
        financial = financial_by_code.get(code, {})
        news = news_by_code.get(code, {})
        score_reasons = split_reason_text(score.get("主要理由"))
        tech_reasons = split_reason_text(tech.get("主要理由"))
        risk_reasons = score_reasons or tech_reasons
        stock = {
            "code": clean_json_value(code),
            "name": clean_json_value(tech.get("股票名稱") or score.get("公司名稱")),
            "market": clean_json_value(tech.get("市場")),
            "category": clean_json_value(tech.get("投資分類")),
            "price": clean_json_value(tech.get("最新價格")),
            "unrealizedPnlRate": None,
            "totalScore": clean_json_value(score.get("總分")),
            "status": clean_json_value(score.get("第二階段狀態") or tech.get("系統狀態")),
            "technicalStatus": clean_json_value(tech.get("系統狀態")),
            "riskReasons": risk_reasons,
            "manualCheck": clean_json_value(score.get("需要人工確認") or tech.get("是否需要人工確認")),
            "updatedAt": clean_json_value(tech.get("更新日期")),
            "sourceStatus": clean_json_value(tech.get("API更新狀態")),
            "sourceLabel": public_source(tech.get("資料來源"), publish_raw_paths),
            "dataQualityStatus": "資料完整",
            "technical": {
                "open": clean_json_value(tech.get("開盤價")),
                "high": clean_json_value(tech.get("最高價")),
                "low": clean_json_value(tech.get("最低價")),
                "close": clean_json_value(tech.get("收盤價")),
                "volume": clean_json_value(tech.get("成交量")),
                "ma5": clean_json_value(tech.get("MA5")),
                "ma10": clean_json_value(tech.get("MA10")),
                "ma20": clean_json_value(tech.get("MA20")),
                "ma60": clean_json_value(tech.get("MA60")),
                "ma120": clean_json_value(tech.get("MA120")),
                "ma240": clean_json_value(tech.get("MA240")),
                "k": clean_json_value(tech.get("K")),
                "d": clean_json_value(tech.get("D")),
                "j": clean_json_value(tech.get("J")),
                "macd": clean_json_value(tech.get("MACD")),
                "macdSignal": clean_json_value(tech.get("MACD Signal")),
                "macdHistogram": clean_json_value(tech.get("MACD Histogram")),
                "rsi": clean_json_value(tech.get("RSI14")),
                "return20d": clean_json_value(tech.get("20日漲跌幅")),
                "return60d": clean_json_value(tech.get("60日漲跌幅")),
                "return120d": clean_json_value(tech.get("120日漲跌幅")),
            },
            "scores": {
                "technical": clean_json_value(score.get("技術面分數")),
                "fundamental": clean_json_value(score.get("基本面分數")),
                "chip": clean_json_value(score.get("籌碼面分數")),
                "news": clean_json_value(score.get("新聞與產業分數")),
                "total": clean_json_value(score.get("總分")),
            },
            "details": {
                "technical": pick_public(
                    tech,
                    ["開盤價", "最高價", "最低價", "收盤價", "成交量", "MA5", "MA10", "MA20", "MA60", "MA120", "MA240", "K", "D", "J", "MACD", "MACD Signal", "MACD Histogram", "RSI14", "20日漲跌幅", "60日漲跌幅", "120日漲跌幅", "跌破MA20", "跌破MA60", "跌破MA120", "放量下跌", "高檔爆量長黑", "系統狀態", "主要理由"],
                    publish_raw_paths,
                ),
                "revenue": pick_public(
                    revenue,
                    ["資料年月", "當月營收", "月增率", "年增率", "累計年增率", "連續幾個月年增", "連續幾個月年減", "基本面警訊", "基本面狀態"],
                    publish_raw_paths,
                ),
                "chip": pick_public(
                    chip,
                    ["日期", "外資買賣超", "投信買賣超", "自營商買賣超", "三大法人合計買賣超", "近5日外資買賣超", "近5日投信買賣超", "近20日三大法人買賣超", "是否連續賣超", "籌碼警訊", "籌碼狀態"],
                    publish_raw_paths,
                ),
                "financial": pick_public(
                    financial,
                    ["EPS", "近四季EPS", "毛利率", "營業利益率", "淨利率", "本益比", "股價淨值比", "殖利率", "財報更新日期", "財報警訊", "財報狀態", "公司下修展望"],
                    publish_raw_paths,
                ),
                "news": pick_public(
                    news,
                    ["公司新聞標題", "新聞日期", "新聞來源", "新聞連結", "新聞摘要", "正面/中性/負面", "是否重大利空", "產業趨勢摘要", "新聞風險警訊", "新聞風險狀態"],
                    publish_raw_paths,
                ),
                "score": pick_public(
                    score,
                    ["技術面分數", "基本面分數", "籌碼面分數", "新聞與產業分數", "總分", "第二階段狀態", "主要理由", "需要人工確認", "不可自動下單", "不可直接執行交易", "資料來源與日期"],
                    publish_raw_paths,
                ),
            },
        }
        latest_price = tech.get("最新價格")
        if latest_price in (None, "", "資料不足"):
            stock["dataQualityStatus"] = "API 失敗" if str(stock.get("sourceStatus") or "").startswith("失敗：") else "部分資料不足"
        elif str(stock.get("manualCheck")) == "是":
            stock["dataQualityStatus"] = "需人工確認"
        else:
            stock["dataQualityStatus"] = "資料完整"
        if show_pnl_rate:
            private_pos = private_positions.get(code, {})
            shares_value = private_pos.get("shares")
            avg_cost_value = private_pos.get("averageCost")
            shares_num = float(shares_value) if isinstance(shares_value, (int, float)) else None
            avg_cost_num = float(avg_cost_value) if isinstance(avg_cost_value, (int, float)) else None
            latest_num = float(latest_price) if isinstance(latest_price, (int, float)) else None
            if shares_num is None or shares_num <= 0:
                stock["unrealizedPnlRate"] = "非持倉"
            elif latest_num is None:
                stock["unrealizedPnlRate"] = "缺價格"
            elif avg_cost_num is None or avg_cost_num == 0:
                stock["unrealizedPnlRate"] = "缺成本"
            else:
                stock["unrealizedPnlRate"] = round((latest_num - avg_cost_num) / avg_cost_num * 100, 2)
        else:
            stock["unrealizedPnlRate"] = "未公開"
        stocks.append(stock)

    def numeric(value: Any) -> float:
        return float(value) if isinstance(value, (int, float)) else 0.0

    total_market_value = sum(numeric(stock.get("marketValue")) for stock in stocks)
    total_pnl = sum(numeric(stock.get("unrealizedPnl")) for stock in stocks)
    invested = total_market_value - total_pnl
    summary = {
        "asOf": run_date,
        "totalMarketValue": round(total_market_value, 2) if publish_position_amounts else None,
        "totalUnrealizedPnl": round(total_pnl, 2) if publish_position_amounts else None,
        "overallReturnRate": round(total_pnl / invested, 4) if publish_position_amounts and invested else None,
        "exitCount": sum(1 for stock in stocks if stock.get("status") == "出場警訊"),
        "reduceCount": sum(1 for stock in stocks if stock.get("status") == "減碼警訊"),
        "manualCheckCount": sum(1 for stock in stocks if stock.get("manualCheck") == "是"),
        "stockCount": len(stocks),
        "averageScore": round(sum(float(s["totalScore"]) for s in stocks if isinstance(s.get("totalScore"), (int, float))) / max(sum(1 for s in stocks if isinstance(s.get("totalScore"), (int, float))), 1), 1),
        "positionAmountsPublished": publish_position_amounts,
        "viewMode": view_settings.get("viewMode", "family"),
    }
    payload = {
        "generatedAt": datetime.now(ZoneInfo("Asia/Taipei")).isoformat(timespec="seconds"),
        "privacy": {
            "publicDashboard": True,
            "positionAmountsPublished": publish_position_amounts,
            "sharesPublished": publish_shares,
            "averageCostPublished": publish_average_cost,
            "rawSourcePathsPublished": publish_raw_paths,
        },
        "summary": summary,
        "stocks": stocks,
        "disclaimer": "本系統僅供投資追蹤與風險提示，不構成買賣建議，不得自動下單，所有決策需人工確認。",
    }

    json_path = docs_data / "stocks.json"
    js_path = docs_data / "stocks-data.js"
    json_text = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    json_path.write_text(json_text, encoding="utf-8")
    js_path.write_text("window.STOCK_DASHBOARD_DATA = " + json_text + ";\n", encoding="utf-8")
    return json_path


def sheet_records(wb: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [norm_text(ws.cell(2, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row in range(3, ws.max_row + 1):
        record: dict[str, Any] = {}
        has_value = False
        for col, header in enumerate(headers, 1):
            if not header:
                continue
            value = ws.cell(row, col).value
            if value not in (None, ""):
                has_value = True
            record[header] = value
        if has_value:
            rows.append(record)
    return rows


def export_dashboard_from_workbook(report_path: Path, public_options: dict[str, Any] | None = None) -> Path:
    wb = load_workbook(report_path, data_only=True)
    result_rows = sheet_records(wb, "04_自動更新結果")
    revenue_rows = sheet_records(wb, "11_月營收追蹤")
    chip_rows = sheet_records(wb, "12_法人籌碼追蹤")
    financial_rows = sheet_records(wb, "13_財報估值追蹤")
    news_rows = sheet_records(wb, "14_新聞風險追蹤")
    score_rows = sheet_records(wb, "15_第二階段綜合評分")
    run_date = datetime.now(ZoneInfo("Asia/Taipei")).date().isoformat()
    return export_dashboard_json(result_rows, revenue_rows, chip_rows, financial_rows, news_rows, score_rows, run_date, public_options)


def run(dry_run: bool = False) -> tuple[Path, Path, Path]:
    config = load_config()
    ensure_dirs(config)
    timezone = config.get("timezone", "Asia/Taipei")
    run_date = datetime.now(ZoneInfo(timezone)).date().isoformat()
    log_path = ROOT / config["logs_folder"] / f"update_log_{run_date}.txt"
    logger = setup_logger(log_path)
    input_excel = ROOT / config["input_excel"]
    output_excel = ROOT / config["reports_folder"] / f"output_report_stage2_{run_date}.xlsx"
    daily_report = ROOT / config["reports_folder"] / f"daily_report_stage2_{run_date}.md"

    logger.info("Start stock tracker run. input=%s output=%s dry_run=%s", input_excel, output_excel, dry_run)
    holdings = read_holdings(input_excel, config["manual_sheet"], config["manual_columns"])
    tracked = [h for h in holdings if h.tracking == "Y"]
    logger.info("Read holdings=%s tracked=%s", len(holdings), len(tracked))

    timeout = int(config.get("request_timeout_seconds", 20))
    start_date = date.today() - timedelta(days=int(config.get("history_calendar_days", 420)))
    fundamental_start_date = date.today() - timedelta(days=int(config.get("fundamental_calendar_days", 900)))
    chip_start_date = date.today() - timedelta(days=int(config.get("chip_calendar_days", 45)))
    news_start_date = date.today() - timedelta(days=int(config.get("news_calendar_days", 14)))
    end_date = date.today()
    result_rows: list[dict[str, Any]] = []
    revenue_rows: list[dict[str, Any]] = []
    chip_rows: list[dict[str, Any]] = []
    financial_rows: list[dict[str, Any]] = []
    news_rows: list[dict[str, Any]] = []
    score_rows: list[dict[str, Any]] = []

    for holding in holdings:
        tech_row: dict[str, Any]
        if holding.tracking != "Y":
            logger.info("%s %s skipped because tracking=%s", holding.stock_id, holding.stock_name, holding.tracking)
            tech_row = build_result_row(holding, None, "未追蹤", "略過：是否追蹤不是 Y", "此股票未啟用追蹤")
            result_rows.append(tech_row)
            rev = build_monthly_revenue_row(holding, None, tech_row, "未追蹤", "略過：是否追蹤不是 Y", "此股票未啟用追蹤")
            chip = build_institutional_row(holding, None, tech_row, "未追蹤", "略過：是否追蹤不是 Y", "此股票未啟用追蹤")
            fin = build_financial_row(holding, None, None, tech_row, "未追蹤", "略過：是否追蹤不是 Y", "此股票未啟用追蹤")
            news = build_news_row(holding, None, "未追蹤", "略過：是否追蹤不是 Y", "此股票未啟用追蹤")
            revenue_rows.append(rev)
            chip_rows.append(chip)
            financial_rows.append(fin)
            news_rows.append(news)
            score_rows.append(build_score_row(holding, tech_row, rev, chip, fin, news))
            continue
        if holding.market not in {"TWSE", "TPEX"}:
            logger.warning("%s unsupported market=%s", holding.stock_id, holding.market)
            tech_row = build_result_row(holding, None, "資料不足", f"失敗：尚未支援市場 {holding.market}", "第一版僅支援 TWSE/TPEx 台股架構")
            result_rows.append(tech_row)
            rev = build_monthly_revenue_row(holding, None, tech_row, "資料不足", "失敗：市場不支援", "第二階段目前僅支援 TWSE/TPEx 台股架構")
            chip = build_institutional_row(holding, None, tech_row, "資料不足", "失敗：市場不支援", "第二階段目前僅支援 TWSE/TPEx 台股架構")
            fin = build_financial_row(holding, None, None, tech_row, "資料不足", "失敗：市場不支援", "第二階段目前僅支援 TWSE/TPEx 台股架構")
            news = build_news_row(holding, None, "資料不足", "失敗：市場不支援", "第二階段目前僅支援 TWSE/TPEx 台股架構")
            revenue_rows.append(rev)
            chip_rows.append(chip)
            financial_rows.append(fin)
            news_rows.append(news)
            score_rows.append(build_score_row(holding, tech_row, rev, chip, fin, news))
            continue
        try:
            if dry_run:
                raise RuntimeError("dry-run 模式未呼叫 API")
            history, raw_path = fetch_finmind_history(holding.stock_id, start_date, end_date, config, timeout, logger)
            if len(history) < 20:
                raise RuntimeError(f"歷史資料筆數不足：{len(history)}")
            enriched = add_indicators(history)
            clean_path = ROOT / config["clean_data_folder"] / f"price_{holding.stock_id}_{run_date}.csv"
            enriched.to_csv(clean_path, index=False, encoding="utf-8-sig")
            logger.info("%s success rows=%s raw=%s clean=%s", holding.stock_id, len(enriched), raw_path, clean_path)
            source = f"FinMind TaiwanStockPrice；raw={raw_path}"
            tech_row = build_result_row(holding, enriched, source, "成功")
            result_rows.append(tech_row)
        except Exception as exc:  # keep one failed stock from stopping the portfolio
            logger.error("%s failed: %s", holding.stock_id, exc)
            tech_row = build_result_row(holding, None, "資料不足", f"失敗：{exc}", str(exc))
            result_rows.append(tech_row)

        try:
            if dry_run:
                raise RuntimeError("dry-run 模式未呼叫 API")
            revenue_df, revenue_raw = fetch_finmind_dataset("TaiwanStockMonthRevenue", holding.stock_id, fundamental_start_date, end_date, config, timeout, logger)
            rev = build_monthly_revenue_row(holding, revenue_df, tech_row, f"FinMind TaiwanStockMonthRevenue；raw={revenue_raw}", "成功")
        except Exception as exc:
            logger.error("%s monthly revenue failed: %s", holding.stock_id, exc)
            rev = build_monthly_revenue_row(holding, None, tech_row, "資料不足", f"失敗：{exc}", str(exc))
        revenue_rows.append(rev)

        try:
            if dry_run:
                raise RuntimeError("dry-run 模式未呼叫 API")
            chip_df, chip_raw = fetch_finmind_dataset("TaiwanStockInstitutionalInvestorsBuySell", holding.stock_id, chip_start_date, end_date, config, timeout, logger)
            chip = build_institutional_row(holding, chip_df, tech_row, f"FinMind TaiwanStockInstitutionalInvestorsBuySell；raw={chip_raw}", "成功")
        except Exception as exc:
            logger.error("%s institutional failed: %s", holding.stock_id, exc)
            chip = build_institutional_row(holding, None, tech_row, "資料不足", f"失敗：{exc}", str(exc))
        chip_rows.append(chip)

        financial_df = None
        per_df = None
        financial_sources: list[str] = []
        financial_status = "成功"
        financial_error = ""
        try:
            if dry_run:
                raise RuntimeError("dry-run 模式未呼叫 API")
            financial_df, financial_raw = fetch_finmind_dataset("TaiwanStockFinancialStatements", holding.stock_id, fundamental_start_date, end_date, config, timeout, logger)
            financial_sources.append(f"FinMind TaiwanStockFinancialStatements；raw={financial_raw}")
        except Exception as exc:
            financial_status = f"部分失敗：財報 {exc}"
            financial_error = str(exc)
            logger.error("%s financial statements failed: %s", holding.stock_id, exc)
        try:
            if dry_run:
                raise RuntimeError("dry-run 模式未呼叫 API")
            per_df, per_raw = fetch_finmind_dataset("TaiwanStockPER", holding.stock_id, start_date, end_date, config, timeout, logger)
            financial_sources.append(f"FinMind TaiwanStockPER；raw={per_raw}")
        except Exception as exc:
            financial_status = f"部分失敗：估值 {exc}" if financial_status == "成功" else financial_status + f"；估值 {exc}"
            logger.error("%s PER failed: %s", holding.stock_id, exc)
        fin = build_financial_row(holding, financial_df, per_df, tech_row, "；".join(financial_sources) if financial_sources else "資料不足", financial_status, financial_error)
        financial_rows.append(fin)

        try:
            if dry_run:
                raise RuntimeError("dry-run 模式未呼叫 API")
            news_df, news_raw = fetch_finmind_dataset("TaiwanStockNews", holding.stock_id, news_start_date, end_date, config, timeout, logger)
            news = build_news_row(holding, news_df, f"FinMind TaiwanStockNews；raw={news_raw}", "成功")
        except Exception as exc:
            logger.error("%s news failed: %s", holding.stock_id, exc)
            news = build_news_row(holding, None, "資料不足", f"失敗：{exc}", str(exc))
        news_rows.append(news)
        score_rows.append(build_score_row(holding, tech_row, rev, chip, fin, news))

    shutil.copy2(input_excel, output_excel)
    wb = load_workbook(output_excel)
    if config["output_sheet"] not in wb.sheetnames:
        wb.create_sheet(config["output_sheet"])
    if config["source_sheet"] not in wb.sheetnames:
        wb.create_sheet(config["source_sheet"])
    stage2_sheets = config.get("stage2_sheets") or {
        "monthly_revenue": "11_月營收追蹤",
        "institutional": "12_法人籌碼追蹤",
        "financial": "13_財報估值追蹤",
        "news": "14_新聞風險追蹤",
        "score": "15_第二階段綜合評分",
    }
    for sheet_name in stage2_sheets.values():
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

    headers = [
        "股票代號", "股票名稱", "市場", "投資分類", "持有股數", "成本均價", "最新價格", "市值",
        "未實現損益", "未實現損益率", "更新日期", "開盤價", "最高價", "最低價", "收盤價", "成交量",
        "MA5", "MA10", "MA20", "MA60", "MA120", "MA240", "K", "D", "J", "MACD", "MACD Signal",
        "MACD Histogram", "RSI14", "20日漲跌幅", "60日漲跌幅", "120日漲跌幅", "跌破MA20",
        "跌破MA60", "跌破MA120", "放量下跌", "高檔爆量長黑", "系統狀態", "主要理由",
        "是否需要人工確認", "資料來源", "API更新狀態",
    ]
    clear_and_write(wb[config["output_sheet"]], headers, result_rows)
    clear_and_write(
        wb[stage2_sheets.get("monthly_revenue", "11_月營收追蹤")],
        ["公司代號", "公司名稱", "資料年月", "當月營收", "上月營收", "去年同月營收", "月增率", "年增率", "累計營收", "去年累計營收", "累計年增率", "連續幾個月年增", "連續幾個月年減", "基本面警訊", "基本面狀態", "資料來源", "API更新狀態"],
        revenue_rows,
    )
    clear_and_write(
        wb[stage2_sheets.get("institutional", "12_法人籌碼追蹤")],
        ["日期", "公司代號", "公司名稱", "外資買賣超", "投信買賣超", "自營商買賣超", "三大法人合計買賣超", "近5日外資買賣超", "近5日投信買賣超", "近20日三大法人買賣超", "是否連續賣超", "籌碼警訊", "籌碼狀態", "資料來源", "API更新狀態"],
        chip_rows,
    )
    clear_and_write(
        wb[stage2_sheets.get("financial", "13_財報估值追蹤")],
        ["公司代號", "公司名稱", "EPS", "近四季EPS", "毛利率", "營業利益率", "淨利率", "本益比", "股價淨值比", "殖利率", "財報更新日期", "財報警訊", "財報狀態", "公司下修展望", "資料來源", "API更新狀態"],
        financial_rows,
    )
    clear_and_write(
        wb[stage2_sheets.get("news", "14_新聞風險追蹤")],
        ["公司代號", "公司名稱", "公司新聞標題", "新聞日期", "新聞來源", "新聞連結", "新聞摘要", "正面/中性/負面", "是否重大利空", "產業趨勢摘要", "新聞風險警訊", "新聞風險狀態", "資料來源", "API更新狀態"],
        news_rows,
    )
    clear_and_write(
        wb[stage2_sheets.get("score", "15_第二階段綜合評分")],
        ["公司代號", "公司名稱", "技術面分數", "基本面分數", "籌碼面分數", "新聞與產業分數", "總分", "第二階段狀態", "主要理由", "需要人工確認", "不可自動下單", "不可直接執行交易", "資料來源與日期"],
        score_rows,
    )
    write_source_sheet(wb[config["source_sheet"]], run_date, result_rows)
    update_manual_api_columns(wb, config, {r["股票代號"]: r for r in result_rows})
    wb.save(output_excel)
    write_stage2_report(score_rows, revenue_rows, chip_rows, news_rows, daily_report, run_date)
    dashboard_json = export_dashboard_json(
        result_rows,
        revenue_rows,
        chip_rows,
        financial_rows,
        news_rows,
        score_rows,
        run_date,
        config.get("public_dashboard", {}),
    )
    logger.info("Done. output_excel=%s daily_report=%s dashboard_json=%s log=%s", output_excel, daily_report, dashboard_json, log_path)
    return output_excel, daily_report, log_path


def main() -> None:
    parser = argparse.ArgumentParser(description="每日更新 Excel 股票追蹤系統")
    parser.add_argument("--dry-run", action="store_true", help="不呼叫外部 API，只驗證 Excel 回寫、報告與 log 流程")
    parser.add_argument("--export-dashboard-from", help="從既有第二階段 Excel 報表匯出 docs/data/stocks.json，不重新呼叫 API")
    args = parser.parse_args()
    if args.export_dashboard_from:
        config = load_config()
        dashboard_json = export_dashboard_from_workbook(Path(args.export_dashboard_from), config.get("public_dashboard", {}))
        print(f"Dashboard JSON: {dashboard_json}")
        return
    output_excel, daily_report, log_path = run(dry_run=args.dry_run)
    print(f"Excel report: {output_excel}")
    print(f"Daily report: {daily_report}")
    print(f"Log: {log_path}")


if __name__ == "__main__":
    main()
